import json
import glob
import os
import pandas as pd
from collections import defaultdict
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

def export_trivy_to_excel(directory_path="."):
    package_data = defaultdict(lambda: {"total_occurrences": 0, "cves": {}, "affected_images": set()})
    files = glob.glob(os.path.join(directory_path, "*.json"))
    if not files:
        print("No JSON files found in the specified directory.")
        return
    for file in files:
        try:
            with open(file, 'r', encoding='utf-8') as f:
                data = json.load(f)
                artifact_name = data.get('ArtifactName')
                if not artifact_name:
                    artifact_name = os.path.basename(file).replace('.json', '')
                for result in data.get('Results', []):
                    result_class = result.get('Class', '')
                    if result_class in ['config', 'secret', 'license']:
                        continue 
                    pkg_type = result.get('Type', 'Unknown Type')
                    for vuln in result.get('Vulnerabilities', []):
                        pkg_name = vuln.get('PkgName', 'Unknown Library')
                        cve_id = vuln.get('VulnerabilityID', 'Unknown CVE')
                        fixed_version = vuln.get('FixedVersion')
                        status = vuln.get('Status', 'unknown')
                        severity = vuln.get('Severity', 'UNKNOWN')
                        dict_key = (pkg_type, pkg_name)
                        package_data[dict_key]["total_occurrences"] += 1
                        package_data[dict_key]["affected_images"].add(artifact_name)
                        if cve_id not in package_data[dict_key]["cves"]:
                            has_patch = "Yes" if fixed_version else "No"
                            package_data[dict_key]["cves"][cve_id] = {
                                "has_patch": has_patch,
                                "fixed_version": fixed_version if fixed_version else "N/A",
                                "status": status,
                                "severity": severity
                            }
        except Exception as e:
            print(f"Error reading {file}: {e}")
    summary_list = []
    type_group_list = []
    details_list = []
    
    for (pkg_type, pkg_name), data in package_data.items():
        unique_vulns = len(data["cves"])
        affected_images_str = ", ".join(sorted(list(data["affected_images"])))
        summary_list.append({
            "Library": pkg_name,
            "Total Occurrences": data["total_occurrences"],
            "Unique Vulnerabilities": unique_vulns,
            "Affected Images": affected_images_str
        })

        type_group_list.append({
            "Type (Ecosystem)": pkg_type.upper(),
            "Library": pkg_name,
            "Total Occurrences": data["total_occurrences"],
            "Unique Vulnerabilities": unique_vulns,
            "Affected Images": affected_images_str
        })
        
        for cve, info in data["cves"].items():
            details_list.append({
                "Type": pkg_type.upper(),
                "Library": pkg_name,
                "CVE ID": cve,
                "Severity": info["severity"],
                "Patch Available?": info["has_patch"],
                "Fixed Version": info["fixed_version"],
                "Status": info["status"] if info["status"] != 'unknown' else "N/A"
            })
    if summary_list and details_list:
        summary_df = pd.DataFrame(summary_list).sort_values(by="Unique Vulnerabilities", ascending=False)
        type_group_df = pd.DataFrame(type_group_list).sort_values(by=["Type (Ecosystem)", "Unique Vulnerabilities"], ascending=[True, False])
        details_df = pd.DataFrame(details_list).sort_values(by=["Library", "Severity"])
    else:
        print("No vulnerabilities found to export.")
        return

    output_filename = "Trivy_Vulnerability_Analysis.xlsx"
    with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
        summary_df.to_excel(writer, sheet_name="Library Summary", index=False)
        type_group_df.to_excel(writer, sheet_name="Libraries by Type", index=False)
        details_df.to_excel(writer, sheet_name="Vulnerability Details", index=False)      
        header_fill = PatternFill(start_color="2B3A42", end_color="2B3A42", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        thin_border = Border(left=Side(style='thin', color='E0E0E0'), 
                             right=Side(style='thin', color='E0E0E0'), 
                             top=Side(style='thin', color='E0E0E0'), 
                             bottom=Side(style='thin', color='E0E0E0'))
        patch_yes_fill = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")
        patch_yes_font = Font(color="137333", bold=True)
        patch_no_fill = PatternFill(start_color="FCE8E6", end_color="FCE8E6", fill_type="solid")
        patch_no_font = Font(color="C5221F", bold=True)
        for sheet_name in writer.sheets:
            worksheet = writer.sheets[sheet_name]
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                cell.border = thin_border
            for col in worksheet.columns:
                max_length = 0
                column_letter = col[0].column_letter
                
                is_affected_images_col = (worksheet[column_letter + '1'].value == "Affected Images")
                
                for cell in col:
                    cell.border = thin_border
                    if cell.row > 1:
                         cell.alignment = left_align
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                if is_affected_images_col:
                     worksheet.column_dimensions[column_letter].width = min(max_length + 4, 50)
                else:
                     worksheet.column_dimensions[column_letter].width = (max_length + 4)
            worksheet.freeze_panes = "A2"
            if sheet_name == "Vulnerability Details":
                patch_col_idx = details_df.columns.get_loc("Patch Available?") + 1
                for row in range(2, worksheet.max_row + 1):
                    cell = worksheet.cell(row=row, column=patch_col_idx)
                    cell.alignment = center_align
                    if cell.value == "Yes":
                        cell.fill = patch_yes_fill
                        cell.font = patch_yes_font
                    elif cell.value == "No":
                        cell.fill = patch_no_fill
                        cell.font = patch_no_font

    print("Excel file generated successfully.")

if __name__ == "__main__":
    export_trivy_to_excel(".")